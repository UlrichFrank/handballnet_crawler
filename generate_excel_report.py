#!/usr/bin/env python3
"""
Generate Excel report from game-centric handball_games.json
Shows for each team: HOME and AWAY games with all player statistics
"""

import json
import openpyxl
import sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from collections import OrderedDict
from pathlib import Path

def load_config(config_file: str = "config.json"):
    """Load config from specified file"""
    config_path = Path(config_file)
    if not config_path.exists():
        config_path = Path(__file__).parent / "config" / config_file
    
    if not config_path.exists():
        print(f"❌ Config file not found: {config_path}")
        sys.exit(1)
    
    with open(config_path, 'r') as f:
        return json.load(f)

def get_league_config(config_file: str = "config.json", league_name_arg=None):
    """Get league configuration"""
    config = load_config(config_file)
    
    if league_name_arg:
        for league in config['leagues']:
            if league['name'] == league_name_arg:
                return league
        print(f"Error: League '{league_name_arg}' not found in config")
        sys.exit(1)
    else:
        # Use first league as default
        return config['leagues'][0]

def load_games_data(data_folder):
    """Load game-centric data from frontend/public/data/{data_folder}/*.json files"""
    # Load all yyyymmdd.json files for this liga
    data_dir = Path('frontend/public/data') / data_folder
    
    if not data_dir.exists():
        print(f"   ❌ Data directory not found: {data_dir}")
        raise FileNotFoundError(f"Data directory not found: {data_dir}")
    
    # Collect all games from all date-based files
    all_games = []
    date_files = sorted(list(data_dir.glob('*.json')))
    
    if not date_files:
        print(f"   ❌ No spieltag files found in {data_dir}")
        raise FileNotFoundError(f"No spieltag files found in {data_dir}")
    
    print(f"   📂 Loading {len(date_files)} Spieltag(e) from {data_dir}")
    
    for date_file in date_files:
        with open(date_file, 'r') as f:
            data = json.load(f)
        games = data.get('games', [])
        all_games.extend(games)
        print(f"      ✅ {date_file.name}: {len(games)} games")
    
    # Return combined data
    return {'games': all_games}

def create_report():
    # Parse command line arguments properly
    config_file = "config.json"  # Default
    league_name_arg = None
    
    # Manual parsing to handle --config flag
    i = 1
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "--config" and i + 1 < len(sys.argv):
            config_file = sys.argv[i + 1]
            i += 2  # Skip both --config and its value
        else:
            league_name_arg = arg
            i += 1
    
    # Get leagues to process
    config = load_config(config_file)
    
    # If league_name argument provided, process only that league
    if league_name_arg:
        leagues_to_process = [get_league_config(config_file, league_name_arg)]
    else:
        # Process all configured leagues
        leagues_to_process = config['leagues']
    
    # Process each league
    for league_config in leagues_to_process:
        league_name = league_config['name']
        data_folder = league_config['name']
        output_file = f"output/{league_name}.xlsx"
        
        print(f"\n📊 Generiere Excel Report für: {league_config['display_name']}")
        print(f"   Lade Spieldaten...")
        
        try:
            data = load_games_data(data_folder)
        except FileNotFoundError:
            print(f"   ⚠️  JSON-Datei nicht gefunden für: {league_config['display_name']}")
            continue
        
        games = data['games']
        
        # Collect all teams and their games (home and away)
        team_games = {}
        
        for game in games:
            home_team = game['home']['team_name']
            away_team = game['away']['team_name']
            game_id = game['game_id']
            order = game.get('order', 0)
            date = game.get('date', 'Unknown')
            
            if not home_team or not away_team:
                continue
            
            # Calculate score from player data
            home_goals = sum(p['goals'] for p in game['home']['players'])
            away_goals = sum(p['goals'] for p in game['away']['players'])
            score = f"{home_goals}:{away_goals}"
            
            # Get graphic path if available
            graphic_path = game.get('graphic_path')
            
            # Home game
            if home_team not in team_games:
                team_games[home_team] = OrderedDict()
            team_games[home_team][game_id] = {
                'order': order,
                'date': date,
                'score': score,
                'opponent': away_team,
                'is_home': True,
                'players': game['home']['players'],
                'graphic_path': graphic_path
            }
            
            # Away game
            if away_team not in team_games:
                team_games[away_team] = OrderedDict()
            team_games[away_team][game_id] = {
                'order': order,
                'date': date,
                'score': score,
                'opponent': home_team,
                'is_home': False,
                'players': game['away']['players'],
                'graphic_path': graphic_path
            }
        
        print(f"   📋 {len(team_games)} Teams gefunden")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        h_font = Font(color="FFFFFF", bold=True, size=10)
        s_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        s_font = Font(bold=True, size=9)
        p_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        p_fill_alt = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        p_font = Font(bold=True)
        c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=10)
        
        # Alternating row fills for players (lighter colors)
        row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        row_fill_2 = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray
        
        labels = ["Tore", "7m Vers.", "7m Tore", "2-Min", "Gelb", "Rot", "Blau"]
        
        # Sort teams alphabetically, but preserve game order from Spielplan
        for tidx, (team_name, games_dict) in enumerate(sorted(team_games.items()), 1):
            print(f"   [{tidx}/{len(team_games)}] {team_name}...")
            
            sname = team_name.replace('/', '-').replace('?', '').replace('[', '').replace(']', '')[:31]
            ws = wb.create_sheet(title=sname)
            
            all_players_set = set()
            for game_data in games_dict.values():
                for player in game_data['players']:
                    all_players_set.add(player['name'])
            
            players = sorted(list(all_players_set))
            
            # Sort games by order from Spielplan
            sorted_games = sorted(
                games_dict.items(),
                key=lambda x: x[1]['order']
            )
            
            print(f"      -> {len(players)} Spieler, {len(games_dict)} Spiele (Heim + Auswärts)")
            
            col = 2
            for game_id, game_data in sorted_games:
                order = game_data['order']
                date = game_data.get('date', 'Unknown')
                score = game_data.get('score', 'Unknown')
                
                # Show teams in correct order: home vs away
                # But ONLY show icon for the current team (the one this tab is for)
                if game_data['is_home']:
                    # Home game: our team at home (show icon for our team)
                    header = f"{date}\n🏠 {team_name}\nvs\n{game_data['opponent']}\n{score}"
                else:
                    # Away game: opponent at home, our team away (show icon for our team)
                    header = f"{date}\n{game_data['opponent']}\nvs\n🏃 {team_name}\n{score}"
                
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 6)
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = c_align
                cell.border = border
                
                col += 7
            
            col = 2
            for _ in sorted_games:
                for label in labels:
                    cell = ws.cell(row=2, column=col)
                    cell.value = label
                    cell.font = s_font
                    cell.fill = s_fill
                    cell.alignment = c_align
                    cell.border = border
                    col += 1
            
            # Add summary headers
            summary_labels = ["Tore\nGesamt", "7m\nVers.", "7m\nTore", "2-Min\nGesamt", "Gelb", "Rot", "Blau"]
            for label in summary_labels:
                cell = ws.cell(row=2, column=col)
                cell.value = label
                cell.font = s_font
                cell.fill = s_fill
                cell.alignment = c_align
                cell.border = border
                col += 1
            
            a1 = ws.cell(row=1, column=1)
            a1.value = "Match"
            a1.font = h_font
            a1.fill = h_fill
            a1.alignment = c_align
            a1.border = border
            
            a2 = ws.cell(row=2, column=1)
            a2.value = "Player"
            a2.font = s_font
            a2.fill = s_fill
            a2.alignment = c_align
            a2.border = border
            
            game_totals = {}
            player_game_stats = {}  # Track per-player, per-game stats for summary
            
            for prow, player_name in enumerate(players, start=3):
                # Alternating row colors
                is_even_row = (prow - 3) % 2 == 0  # 0-indexed row from player start
                row_fill = row_fill_1 if is_even_row else row_fill_2
                
                ca = ws.cell(row=prow, column=1)
                ca.value = player_name
                ca.font = p_font
                ca.fill = row_fill
                ca.alignment = l_align
                ca.border = border
                
                col = 2
                game_idx = 0
                player_all_stats = [0, 0, 0, 0, 0, 0, 0]  # Total stats for this player across all games
                
                for game_id, game_data in sorted_games:
                    if game_idx not in game_totals:
                        game_totals[game_idx] = [0, 0, 0, 0, 0, 0, 0]
                    
                    player_stats = None
                    for player in game_data['players']:
                        if player['name'] == player_name:
                            player_stats = player
                            break
                    
                    if player_stats:
                        stats = [
                            player_stats['goals'],
                            player_stats.get('seven_meters', 0),
                            player_stats.get('seven_meters_goals', 0),
                            player_stats['two_min_penalties'],
                            player_stats['yellow_cards'],
                            player_stats['red_cards'],
                            player_stats['blue_cards']
                        ]
                        for i, val in enumerate(stats):
                            game_totals[game_idx][i] += val
                            player_all_stats[i] += val
                    else:
                        stats = [0, 0, 0, 0, 0, 0, 0]
                    
                    for idx, stat_val in enumerate(stats):
                        cell = ws.cell(row=prow, column=col)
                        # Tore (idx=0): "-" only if not participated (player_stats is None)
                        if idx == 0:
                            cell.value = stat_val if player_stats else "-"
                        # 7m Tore (idx=2): "-" only if no 7m Vers. attempt (stats[1] == 0)
                        elif idx == 2:
                            cell.value = stat_val if stats[1] > 0 else "-"
                        # 7m Vers., 2-Min, Gelb, Rot, Blau (idx=1,3,4,5,6): "-" if 0
                        elif idx in [1, 3, 4, 5, 6]:
                            cell.value = stat_val if stat_val > 0 else "-"
                        else:
                            cell.value = stat_val
                        cell.alignment = c_align
                        cell.border = border
                        cell.fill = row_fill
                        col += 1
                    
                    game_idx += 1
                
                # Add summary columns for this player
                for idx, stat_val in enumerate(player_all_stats):
                    cell = ws.cell(row=prow, column=col)
                    # Tore (idx=0): Always show number (0 instead of "-")
                    if idx == 0:
                        cell.value = stat_val
                    # 7m Tore (idx=2): "-" only if no 7m Vers. attempts (player_all_stats[1] == 0)
                    elif idx == 2:
                        cell.value = stat_val if player_all_stats[1] > 0 else "-"
                    # 7m Vers., 2-Min, Gelb, Rot, Blau (idx=1,3,4,5,6): "-" if 0
                    elif idx in [1, 3, 4, 5, 6]:
                        cell.value = stat_val if stat_val > 0 else "-"
                    else:
                        cell.value = stat_val
                    cell.alignment = c_align
                    cell.border = border
                    cell.fill = row_fill
                    cell.font = Font(bold=True)
                    col += 1
            
            totals_row = len(players) + 3
            
            ta = ws.cell(row=totals_row, column=1)
            ta.value = "GESAMT"
            ta.font = total_font
            ta.fill = total_fill
            ta.alignment = l_align
            ta.border = border
            
            col = 2
            for game_idx in range(len(games_dict)):
                stats = game_totals[game_idx]
                for idx, stat_val in enumerate(stats):
                    cell = ws.cell(row=totals_row, column=col)
                    # Tore (idx=0): "-" if no goals scored
                    if idx == 0:
                        cell.value = stat_val if stat_val > 0 else "-"
                    # 7m Tore (idx=2): "-" only if no 7m Vers. attempts (stats[1] == 0)
                    elif idx == 2:
                        cell.value = stat_val if stats[1] > 0 else "-"
                    # 7m Vers., 2-Min, Gelb, Rot, Blau (idx=1,3,4,5,6): "-" if 0
                    elif idx in [1, 3, 4, 5, 6]:
                        cell.value = stat_val if stat_val > 0 else "-"
                    else:
                        cell.value = stat_val
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.alignment = c_align
                    cell.border = border
                    col += 1
            
            # Add totals for summary columns
            summary_totals = [0, 0, 0, 0, 0, 0, 0]
            for game_idx in range(len(games_dict)):
                stats = game_totals[game_idx]
                for i, stat_val in enumerate(stats):
                    summary_totals[i] += stat_val
            
            for stat_val in summary_totals:
                cell = ws.cell(row=totals_row, column=col)
                cell.value = stat_val if stat_val > 0 else "-"
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = c_align
                cell.border = border
                col += 1
            
            # Embed graphics under GESAMT row
            current_graphic_row = totals_row + 2  # Leave one blank row
            col = 2
            for game_id, game_data in sorted_games:
                graphic_path = game_data.get('graphic_path')
                
                if graphic_path and Path(graphic_path).exists():
                    try:
                        # Merge cells HORIZONTALLY only (7 columns per game, 1 row)
                        ws.merge_cells(start_row=current_graphic_row, start_column=col, 
                                      end_row=current_graphic_row, end_column=col + 6)
                        
                        # Set row height to accommodate the image
                        # Grafik: 560px wide, 140px high (4:1 ratio)
                        # Excel: 1 point ≈ 1.33 pixels
                        # 140 pixels ≈ 105 points
                        ws.row_dimensions[current_graphic_row].height = 105
                        
                        # Insert image - full width, height proportional
                        img = XLImage(graphic_path)
                        img.width = 560   # 7 Spalten × 80 pixels
                        img.height = 140  # Proportional: 560 × (4/16)
                        
                        ws.add_image(img, f'{openpyxl.utils.get_column_letter(col)}{current_graphic_row}')
                        
                    except Exception as e:
                        print(f"       ⚠️  Grafik-Einbettung fehlgeschlagen: {str(e)[:50]}")
                
                col += 7
            
            ws.column_dimensions['A'].width = 25
            for c in range(2, col):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 10
            
            ws.row_dimensions[1].height = 80
            ws.row_dimensions[2].height = 20
            
            # Freeze panes: column A and row 2
            ws.freeze_panes = 'B3'
        
        wb.save(output_file)
        print(f"   ✅ Gespeichert: {output_file}")
    
    print(f"\n✅ Alle Excel Reports erstellt")

if __name__ == '__main__':
    create_report()
